import json
import openpyxl

# Klasser
class Kunde:
    def __init__(
        self,
        fornavn: str,
        etternavn: str,
        tlf: int,
        epost: str,
        adresse: str
    ):
        self.fornavn = fornavn
        self.etternavn = etternavn
        self.tlf = tlf
        self.epost = epost
        self.adresse = adresse

    def eksporter_bruker(
        self,
        filnavn
    ):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet["I1"] = "fornavn:"
        sheet["J1"] = "etternavn:"
        sheet["K1"] = "tlf:"
        sheet["L1"] = "epost:"
        sheet["M1"] = "adresse:"
        sheet["I2"] = self.fornavn.title()
        sheet["J2"] = self.etternavn.title()
        sheet["K2"] = self.tlf
        sheet["L2"] = self.epost
        sheet["M2"] = self.adresse

        workbook.save(filnavn + ".xlsx")


class Billett:
    def __init__(
        self,
        forestilling: str,
        dato: str,
        honnor: int,
        voksen: int,
        student: int,
        barn: int,
        standarpris=300,
    ):
        self.forestilling = forestilling
        self.dato = dato
        self.voksen = voksen
        self.honnor = honnor
        self.student = student
        self.barn = barn
        self.standarpris = standarpris

    def beregn_pris(
        self
    ):
        totalsum = 0
        pris = self.standarpris
        totalsum += (pris - ((pris / 100) * 30)) * self.honnor
        totalsum += pris * self.voksen
        totalsum += (pris - ((pris / 100) * 20)) * self.student
        totalsum += (pris - ((pris / 100) * 50)) * self.barn
        return totalsum

    def eksporter_bestilling(
        self,
        filnavn
    ):
        workbook = openpyxl.load_workbook(filnavn + ".xlsx")
        sheet = workbook.active
        sheet["A1"] = "forestillinger:"
        sheet["B1"] = "honnørbilletter:"
        sheet["C1"] = "studentbilletter:"
        sheet["D1"] = "barnebilletter:"
        sheet["E1"] = "voksenbilletter:"
        sheet["F1"] = "priser:"
        sheet["G1"] = "datoer:"
        sheet["H1"] = "bestillingsnummer"
        print('H2', sheet["H2"].value)
        bestillingsnummer = sheet["H2"].value
        print('bestillingsnummer', bestillingsnummer)

        if bestillingsnummer == None:
            bestillingsnummer = 2
        sheet["H2"] = bestillingsnummer + 1
        print('h2', sheet["H2"].value)
        workbook.save(filnavn + ".xlsx")
        sheet["A" + str(sheet["H2"].value)] = self.forestilling
        sheet["B" + str(sheet["H2"].value)] = self.honnor
        sheet["C" + str(sheet["H2"].value)] = self.student
        sheet["D" + str(sheet["H2"].value)] = self.barn
        sheet["E" + str(sheet["H2"].value)] = self.voksen
        sheet["F" + str(sheet["H2"].value)] = self.beregn_pris()
        sheet["g" + str(sheet["H2"].value)] = self.dato
        workbook.save(filnavn + ".xlsx")

        print('\nKvitteringen er eksportert')

    def vis_kvittering(
        self,
        filnavn,
        dato
    ):
        workbook = openpyxl.load_workbook(filnavn + ".xlsx")
        sheet = workbook.active

        forestillinger = []
        honnorer = []
        studenter = []
        barn = []
        voksne = []
        priser = []
        datoer = []

        for i in sheet["A"]:
            if not (i == sheet["A1"]):
                forestillinger.append(i.value)
        for i in sheet["B"]:
            if not (i == sheet["B1"]):
                honnorer.append(i.value)
        for i in sheet["C"]:
            if not (i == sheet["C1"]):
                studenter.append(i.value)
        for i in sheet["D"]:
            if not (i == sheet["D1"]):
                barn.append(i.value)
        for i in sheet["E"]:
            if not (i == sheet["E1"]):
                voksne.append(i.value)
        for i in sheet["F"]:
            if not (i == sheet["F1"]):
                priser.append(i.value)
        for i in sheet["G"]:
            if not (i == sheet["G1"]):
                datoer.append(i.value)
        spesifike_ting = [
            "\nDu kan finne følgende info: ",
            "-Hvilken forestilling du så",
            "-Hvor mange honørbilletter du bestillte",
            "-Hvor mange studentbilletter du bestillte",
            "-Hvor mange barnebilletter du bestillte",
            "-Hvor mange voksenbilletter du bestillte",
            "-Hva bestillingen din kostet",
        ]
        print(
            f"\nHer er kvitteringen din:\n"
            f"Forestilling: {forestillinger[datoer.index(dato)]}\n"
            f"Honørbilletter: {honnorer[datoer.index(dato)]}\n"
            f"Studentbilletter: {studenter[datoer.index(dato)]}\n"
            f"Barnebilletter: {barn[datoer.index(dato)]}\n"
            f"Voksenbilletter: {voksne[datoer.index(dato)]}\n"
            f"Pris: {priser[datoer.index(dato)]}\n"
            f"Dato: {datoer[datoer.index(dato)]}\n"
        )

        vis_spesifikasjoner = True
        if len(forestillinger) >= 2:
            while True:
                svar = input(
                    '\nVil du se tidligere kvitteringer?'\
                    '\nSvar(J/N): '
                )
                if svar.upper()[0] == "J":
                    svar = input(
                        'Svar "dato" hvis du skal printe en kvittering for en spesifik dato.'
                        '\nSvar "alle" hvis du skal printe alle kvitteringer.'
                        '\nSvar("dato"/"alle"): '
                    )
                    while True:
                        if svar.lower()[:4] == "dato":
                            svar = input(
                                "\nHvilken dato vil du finne kvittering for?"\
                                "\nSvar(dd/mm/åå): "
                            )
                            hvilken_dato = int(svar[0:2])
                            print(hvilken_dato)
                            if hvilken_dato in datoer:
                                print(
                                    f"\nBestilling {datoer.index(hvilken_dato)+1}:\n",
                                    f"Forestilling: {forestillinger[datoer.index(hvilken_dato)]}\n",
                                    f"Honørbilletter: {honnorer[datoer.index(hvilken_dato)]}\n",
                                    f"Studentbilletter: {studenter[datoer.index(hvilken_dato)]}\n",
                                    f"Barnebilletter: {barn[datoer.index(hvilken_dato)]}\n",
                                    f"Voksenbilletter: {voksne[datoer.index(hvilken_dato)]}\n",
                                    f"Pris: {priser[datoer.index(hvilken_dato)]}\n",
                                    f"Dato: {datoer[datoer.index(hvilken_dato)]}\n"
                                )
                                break
                            else:
                                print(f"\nVi fant desverre ingen kvitteringer for datoen {svar}.\n")
                                svar = input("Vil du prøve på nytt?" "\nSvar(J/N): ")
                                if svar.upper()[0] == "J":
                                    print("\nHusk å svare dato/måned/år(dd/mm/åå)")
                                    svar = "dato"
                                if svar.upper()[0] == "N":
                                    print("\nPrøv gjerne på nytt en annen gang.")
                                    break
                        elif svar.lower()[:4] == "alle":
                            if vis_spesifikasjoner:
                                for i in spesifike_ting:
                                    print(i)
                            svar = input("\nLeter du etter noe spesifikt?" "\nSvar(J/N): ")
                            if svar.upper()[0] == "J":
                                svar = input(
                                    '\nSvar "forestilling" for å finne ut hvilken forestilling du så.'
                                    '\nSvar "honør" for å finne ut hvor mange honørbilletter du bestillte.'
                                    '\nSvar "student" for å finne ut hvor mange studentbilletter du bestillte.'
                                    '\nSvar "barn" for å finne ut hvor mange barnebilletter du bestillte.'
                                    '\nSvar "voksen" for å finne ut hvor mange voksenbilletter du bestillte.'
                                    'n\nSvar "pris" for å finne ut hva bestillingen din kostet.'
                                    "\nSvar: "
                                )
                                print()
                                if svar.lower()[:12] == "forestilling":
                                    for i in range(len(forestillinger)):
                                        print(
                                            f"Forestilling: {forestillinger[i]}"
                                            f"\nDato: {datoer[i]}\n"
                                        )
                                if svar.lower()[:5] == "honør":
                                    for i in range(len(honnorer)):
                                        print(
                                            f"Honørbilletter: {honnorer[i]}"
                                            f"\nDato: {datoer[i]}\n"
                                        )
                                if svar.lower()[:7] == "student":
                                    for i in range(len(studenter)):
                                        print(
                                            f"Studentbilletter: {studenter[i]}"
                                            f"\nDato: {datoer[i]}\n"
                                        )
                                if svar.lower()[:4] == "barn":
                                    for i in range(len(barn)):
                                        print(f"Barnebilletter: {barn[i]}" f"\nDato: {datoer[i]}\n")
                                if svar.lower()[:6] == "voksen":
                                    for i in range(len(voksne)):
                                        print(
                                            f"Voksenbilletter: {voksne[i]}" f"\nDato: {datoer[i]}\n"
                                        )
                                if svar.lower()[:4] == "pris":
                                    for i in range(len(priser)):
                                        print(f"Priser: {priser[i]}" f"\nDato: {datoer[i]}\n")
                                break
                            elif svar.upper()[0] == "N":
                                for i in range(len(datoer)):
                                    print(
                                        f"\nBestilling {i+1}:\n"
                                        f"Forestilling: {forestillinger[i]}\n"
                                        f"Honørbilletter: {honnorer[i]}\n"
                                        f"Studentbilletter: {studenter[i]}\n"
                                        f"Barnebilletter: {barn[i]}\n"
                                        f"Voksenbilletter: {voksne[i]}\n"
                                        f"Priser: {priser[i]}\n"
                                        f"Dato: {datoer[i]}\n"
                                    )
                                break
                            else:
                                svar = "alle"
                                vis_spesifikasjoner = False
                                print("\nDu må svare J(ja) eller N(nei).")
                        else:
                            print('\nDu må skrive "dato" eller "alle".\n')
                            svar = input(
                                'Svar "dato" hvis du skal printe en kvittering for en spesifik dato.'
                                '\nSvar "alle" hvis du skal printe alle kvitteringer.'
                                "\nSvar: "
                            )
                    else:
                        print()
                elif svar.upper()[0] == "N":
                    print(
                        "\nPrøv gjerne på nytt en annen gang."
                    )
                    break
                else:
                    print(
                        "\nDu må svare J(ja) eller N(nei)."
                    )


# Variabler
wb = openpyxl.load_workbook("LedigDato.xlsx")
sheet = wb.active
forestilling_bypass = True
forestillinger = ['De Elendige', 'Vildanden']
if sheet["J2"].value == 'ja':
    forestillinger.append('Smalltalk')
pris = 300

#Intro
print(
    'Hei og velkommen til teateret New Wave!\n'\
    '\nFørst vil du være nødt til å registrere en ny bruker eller logge inn på en alleredeeksisterende bruker'\
    '\nDeretter vil du få muligheten til å velge en forestilling ut av følgende valg:'
)
for i in forestillinger:
    print('-',i)
print(
    '\nDa vil du kunne plassere en bestilling på en forestilling for en gyldig dato'\
    '\nTil slutt vil du se kvitteringen for bestillingen din og ha muligheten til å se på kvitteringer for eventuelle tidligere bestillinger.'
)
# Registrering av bruker
epost = input(
    "\nHvilken epostadresse skal brukeren registreres under? "\
    "\n*I fall du er en returnerende kunde, skriv epostadressen du logget inn med under forgje innlogging"\
    '\nSvar("brukernavn@domene.TDL"): '
)
try:
    workbook = openpyxl.load_workbook(epost + ".xlsx")
    sheet = workbook.active

    fornavn = sheet["I2"].value
    etternavn = sheet["J2"].value
    tlf = sheet["K2"].value
    epost = sheet["L2"].value
    adresse = sheet["M2"].value

    print(
        "\nDu er logget inn"
    )


except FileNotFoundError:
    fornavn = input("\nHvilket fornavn skal brukeren registres under?"\
        '\nSvar: '
    )
    etternavn = input("\nHvilket etternavn skal brukeren registres under?"\
        '\nSvar: '
    )
    tlf_temp = input(
        "\nHvilket telefonnummer skal knyttes oppmot brukeren? "\
        '\nSvar(xxx xx xxx): '
    )
    tlf = tlf_temp.replace(' ','')

    adresse = input("\nHvilken adresse skal registreres under brukeren?"
        '\nSvar("Gate", "Husnummer", "By", "Postnummer"): '
    )

    print(
        "\nBrukeren har blitt opprettet"
    )

ny_kunde = Kunde(fornavn, etternavn, tlf, epost, adresse)
ny_kunde.eksporter_bruker(epost)

# Finner ledige datoer
while True:
    forestilling = input(
        "\nHvilken forestilling vil du se?"\
        "\nSvar(liste over forestillinger ligger nær toppen av utskriften): "
    )

    if forestilling.lower() == "de elendige" and "De Elendige" in forestillinger:
        coloumn_to_print = "A"
        coloumn_ledigplass = "C"
        forestilling_bypass = True
    elif forestilling.lower() == "vildanden" and "Vildanden" in forestillinger:
        coloumn_to_print = "D"
        coloumn_ledigplass = "F"
        forestilling_bypass = True
    elif forestilling.lower() == "smalltalk" and "Smalltalk" in forestillinger:
        coloumn_to_print = "G"
        coloumn_ledigplass = "i"

    else:
        forestilling_bypass = False
        print(
            f"\nVi finner ingen visninger for {forestilling}.",
            "\n-Forslag:\n",
            "\n-Skjekk hvilke forestillinger som vises på New Wave.",
            "\n-Skjekk at du har stavet forestillingen rett.",
            '\n-Skjekk for tilsynelatende ubetydelige tegn som "mellomrom", ¨, ~,´ eller ^.'
            "\n-Skjekk at du har rett tastaturspråk."
        )

    if forestilling_bypass:
        wb = openpyxl.load_workbook("LedigDato.xlsx")
        sheet = wb.active


        print(f"\nHer er ledige datoer for {forestilling}:")
        for row in sheet.iter_rows(min_row=2):
            dato1 = row[sheet[coloumn_to_print  + '1'].column - 1].value
            plass1 = row[sheet[coloumn_ledigplass + '1'].column - 1].value
            #Problemer her ja
            if plass1 > 0:
                print(
                    f"\nDato: {dato1}"\
                    f"\nLedige plasser: {plass1}"
                )
        while True:
            try:
                honnor = int(
                    input(
                        "\nHvor mange honnørbilletter skal du ha?"\
                        f"\nPris: {pris - ((pris / 100) * 30)}kr"\
                        "\n*Denne biletten gjelder for personer over 67år"\
                        "\nSvar(heltall): "
                    )
                )
                voksen = int(
                    input(
                        "\nHvor mange voksenbilletter skal du ha?"\
                        f"\nPris: {pris}kr"\
                        "\n*Denne billetten gjelder for personer 10 til 67 år, ikke studenter"\
                        "\nSvar(heltall): "
                    )
                )
                student = int(
                    input(
                        "\nHvor mange studentbilletter skal du ha?"\
                        f"\nPris: {pris - ((pris / 100) * 20)}kr"
                        "\n*Denne billetten gjelder for de som studerer. Dette inkluderer VGS-elever"
                        "\nSvar(heltall): "
                    )
                )
                barn = int(
                    input(
                        "\nHvor mange barnebilletter skal du ha?"\
                        f"\nPris: {pris - ((pris / 100) * 50)}kr"
                        "\n*Denne biletten gjelder for de som er under 10 år"
                        "\nSvar(heltall): "
                    )
                )
                antall_billetter = honnor + voksen + student + barn
                break
            except ValueError:
                print(
                    'Du må skrive inn et gyldig tall.',
                    "\n-Forslag:\n",
                    "\n-Skjekk at du ikke har skrevet tegn slik som (, ), + eller -.",
                    "\n-Skjekk at du ikke har skrevet noen bokstaver."
                    "\n-Skjekk for mellomrom i tall. Ikke gyldig:1 000 000, Gyldig: 1_000_000."
                )
        if antall_billetter > 0:
            dato = input(
                '\nHvilken dato vil du bestille for?'\
                '\nSvar(dd/mm/åå): '
            )
            temp_dato = int(dato[:2])

            if antall_billetter <= sheet[coloumn_ledigplass + str(temp_dato+1)].value:
                temp_plass = sheet[coloumn_ledigplass + str(temp_dato+1)].value
                sheet[coloumn_ledigplass + str(temp_dato+1)] = temp_plass - antall_billetter
                break
            else:
                svar = input(
                    f"De er ingen ledige forestillinger for {forestilling} den {dato}.februar. Vil du velge et annet tidspunkt?"\
                   "\nsvar(J/N): "
                )
                if svar.upper()[0] == "J":
                    True
                else:
                    break
        else:
            print("Du må bestille minst 1 billett")
    else:
        True
wb.save("LedigDato.xlsx")


# Lager billetten
ny_billett = Billett(forestilling, dato, honnor, voksen, student, barn)
ny_billett.eksporter_bestilling(epost)

ny_billett.vis_kvittering(epost, dato)
