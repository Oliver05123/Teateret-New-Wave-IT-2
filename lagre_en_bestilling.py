import openpyxl
import json

# Klasser
class Kunde:
    def __init__(
        self,
        fornavn: str,
        etternavn: str,
        tlf: int,
        epost: str,
        adresse: str
    ):
        self.fornavn = fornavn
        self.etternavn = etternavn
        self.tlf = tlf
        self.epost = epost
        self.adresse = adresse

    def eksporter_bruker(
        self,
        filnavn
    ):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        sheet["I1"] = "fornavn:"
        sheet["J1"] = "etternavn:"
        sheet["K1"] = "tlf:"
        sheet["L1"] = "epost:"
        sheet["M1"] = "adresse:"
        sheet["I2"] = self.fornavn.title()
        sheet["J2"] = self.etternavn.title()
        sheet["K2"] = self.tlf
        sheet["L2"] = self.epost
        sheet["M2"] = self.adresse

        workbook.save(filnavn + ".xlsx")


class Billett:
    def __init__(
        self,
        forestilling: str,
        dato: str,
        honnor: int,
        voksen: int,
        student: int,
        barn: int,
        standarpris=300,
    ):
        self.forestilling = forestilling
        self.dato = dato
        self.voksen = voksen
        self.honnor = honnor
        self.student = student
        self.barn = barn
        self.standarpris = standarpris

    def beregn_pris(
        self
    ):
        totalsum = 0
        pris = self.standarpris
        totalsum += (pris - ((pris / 100) * 30)) * self.honnor
        totalsum += pris * self.voksen
        totalsum += (pris - ((pris / 100) * 20)) * self.student
        totalsum += (pris - ((pris / 100) * 50)) * self.barn
        return totalsum

    def eksporter_bestilling(
        self,
        filnavn
    ):
        workbook = openpyxl.load_workbook("LedigDato.xlsx")
        sheet = workbook.active
        sheet["K2"] = sheet["K2"].value + self.beregn_pris()
        if sheet["K2"].value >= 5_000:
            sheet["J2"] = 'ja'
        else:
            sheet["J2"] = 'nei'
        workbook.save("LedigDato.xlsx")
        workbook = openpyxl.load_workbook(filnavn + ".xlsx")
        sheet = workbook.active
        sheet["A1"] = "forestillinger:"
        sheet["B1"] = "honnørbilletter:"
        sheet["C1"] = "studentbilletter:"
        sheet["D1"] = "barnebilletter:"
        sheet["E1"] = "voksenbilletter:"
        sheet["F1"] = "priser:"
        sheet["G1"] = "datoer:"
        sheet["A2"] = self.forestilling
        sheet["B2"] = self.honnor
        sheet["C2"] = self.student
        sheet["D2"] = self.barn
        sheet["E2"] = self.voksen
        sheet["F2"] = self.beregn_pris()
        sheet["g2"] = self.dato
        workbook.save(filnavn + ".xlsx")


        print('\nKvitteringen er eksportert')

    def vis_kvittering(
        self
    ):
        print(
            f"\nHer er kvitteringen din:\n"
            f"Forestilling: {self.forestilling}\n"
            f"Honørbilletter: {self.honnor}\n"
            f"Studentbilletter: {self.student}\n"
            f"Barnebilletter: {self.barn}\n"
            f"Voksenbilletter: {self.voksen}\n"
            f"Pris: {self.beregn_pris()}\n"
            f"Dato: {self.dato}\n"
        )



# Variabler
wb = openpyxl.load_workbook("LedigDato.xlsx")
sheet = wb.active
forestilling_bypass = True
forestillinger = ['De Elendige', 'Vildanden']
if sheet["J2"].value == 'ja':
    forestillinger.append('Smalltalk')
pris = 300
datoer1 = []
gyldig_dato = False


#Intro
print(
    'Hei og velkommen til teateret New Wave!\n'\
    '\nFørst vil du være nødt til å registrere en ny bruker eller logge inn på en alleredeeksisterende bruker'\
    '\nDeretter vil du få muligheten til å velge en forestilling ut av følgende valg:'
)
for i in forestillinger:
    print('-',i)
print(
    '\nDa vil du kunne plassere en bestilling på en forestilling for en gyldig dato'\
    '\nTil slutt vil du se kvitteringen for bestillingen din og ha muligheten til å se på kvitteringer for eventuelle tidligere bestillinger.'
)
# Registrering av bruker
epost = input(
    "\nHvilken epostadresse skal brukeren registreres under? "\
    "\n*I fall du er en returnerende kunde, skriv epostadressen du logget inn med under forgje innlogging"\
    '\nSvar("brukernavn@domene.TDL"): '
)
try:
    workbook = openpyxl.load_workbook(epost + ".xlsx")
    sheet = workbook.active

    fornavn = sheet["I2"].value
    etternavn = sheet["J2"].value
    tlf = sheet["K2"].value
    epost = sheet["L2"].value
    adresse = sheet["M2"].value

    print(
        "\nDu er logget inn"
    )


except FileNotFoundError:
    fornavn = input("\nHvilket fornavn skal brukeren registres under?"\
        '\nSvar: '
    )
    etternavn = input("\nHvilket etternavn skal brukeren registres under?"\
        '\nSvar: '
    )
    tlf_temp = input(
        "\nHvilket telefonnummer skal knyttes oppmot brukeren? "\
        '\nSvar(xxx xx xxx): '
    )
    tlf = tlf_temp.replace(' ','')

    adresse = input("\nHvilken adresse skal registreres under brukeren?"
        '\nSvar("Gate", "Husnummer", "By", "Postnummer"): '
    )

    print(
        "\nBrukeren har blitt opprettet"
    )

ny_kunde = Kunde(fornavn, etternavn, tlf, epost, adresse)
ny_kunde.eksporter_bruker(epost)

# Finner ledige datoer
while True:
    forestilling = input(
        "\nHvilken forestilling vil du se?"\
        "\nSvar(liste over forestillinger ligger nær toppen av utskriften): "
    )

    if forestilling.lower() == "de elendige" and "De Elendige" in forestillinger:
        coloumn_to_print = "A"
        coloumn_ledigplass = "C"
        forestilling_bypass = True
    elif forestilling.lower() == "vildanden" and "Vildanden" in forestillinger:
        coloumn_to_print = "D"
        coloumn_ledigplass = "F"
        forestilling_bypass = True
    elif forestilling.lower() == "smalltalk" and "Smalltalk" in forestillinger:
        coloumn_to_print = "G"
        coloumn_ledigplass = "i"

    else:
        forestilling_bypass = False
        print(
            f"\nVi finner ingen visninger for {forestilling}.",
            "\n-Forslag:\n",
            "\n-Skjekk hvilke forestillinger som vises på New Wave.",
            "\n-Skjekk at du har stavet forestillingen rett.",
            '\n-Skjekk for tilsynelatende ubetydelige tegn som "mellomrom", ¨, ~,´ eller ^.'
            "\n-Skjekk at du har rett tastaturspråk."
        )

    if forestilling_bypass:
        wb = openpyxl.load_workbook("LedigDato.xlsx")
        sheet = wb.active


        print(f"\nHer er ledige datoer for {forestilling}:")
        for row in sheet.iter_rows(min_row=2):
            dato1 = row[sheet[coloumn_to_print  + '1'].column - 1].value
            plass1 = row[sheet[coloumn_ledigplass + '1'].column - 1].value
            #Problemer her ja
            if plass1 > 0:
                print(
                    f"\nDato: {dato1}"\
                    f"\nLedige plasser: {plass1}"
                )
                datoer1.append(dato1)
        while True:
            try:
                honnor = int(
                    input(
                        "\nHvor mange honnørbilletter skal du ha?"\
                        f"\nPris: {pris - ((pris / 100) * 30)}kr"\
                        "\n*Denne biletten gjelder for personer over 67år"\
                        "\nSvar(Positivt heltall): "
                    )
                )
                voksen = int(
                    input(
                        "\nHvor mange voksenbilletter skal du ha?"\
                        f"\nPris: {pris}kr"\
                        "\n*Denne billetten gjelder for personer 10 til 67 år, ikke studenter"\
                        "\nSvar(Positivt heltall): "
                    )
                )
                student = int(
                    input(
                        "\nHvor mange studentbilletter skal du ha?"\
                        f"\nPris: {pris - ((pris / 100) * 20)}kr"
                        "\n*Denne billetten gjelder for de som studerer. Dette inkluderer VGS-elever"
                        "\nSvar(Positivt heltall): "
                    )
                )
                barn = int(
                    input(
                        "\nHvor mange barnebilletter skal du ha?"\
                        f"\nPris: {pris - ((pris / 100) * 50)}kr"
                        "\n*Denne biletten gjelder for de som er under 10 år"
                        "\nSvar(Positivt heltall): "
                    )
                )
                antall_billetter = honnor + voksen + student + barn
                break
            except ValueError:
                print(
                    'Du må skrive inn et gyldig tall.',
                    "\n-Forslag:\n",
                    "\n-Skjekk at du ikke har skrevet tegn slik som (, ), + eller -.",
                    "\n-Skjekk at du ikke har skrevet noen bokstaver."
                    "\n-Skjekk for mellomrom i tall. Ikke gyldig:1 000 000, Gyldig: 1_000_000."
                )
        if antall_billetter > 0:
            while gyldig_dato == False:
                dato = input(
                    '\nHvilken dato vil du bestille for?'\
                    '\n*Skjekk listen over for ledige datoer'\
                    '\nSvar(dd.mm.åå): '
                )
                temp_dato = int(dato[:2])
                if dato not in datoer1:
                    print(
                        f'\nHei vi finner ingen forestillinger for {dato}'\
                        '\nForslag\n'
                        '\n-Skjekk at datoen er i listen for ledige datoer'
                        '\n-Husk riktig format (dd.mm.åå)'
                        '\n-Du kan ikke ha ugyldige datoer. For eksempel "Gyldig: 01.02.24", "Ikke gydlig: -01.02.24", "Ikke kyldig: 46.02.24'
                    )
                    gyldig_dato = False
                else:
                    gyldig_dato = True
            if antall_billetter <= sheet[coloumn_ledigplass + str(temp_dato+1)].value:
                temp_plass = sheet[coloumn_ledigplass + str(temp_dato+1)].value
                sheet[coloumn_ledigplass + str(temp_dato+1)] = temp_plass - antall_billetter
                break
            else:
                svar = input(
                    f"De er ingen ledige forestillinger for {forestilling} den {dato}.februar. Vil du velge et annet tidspunkt?"\
                   "\nsvar(J/N): "
                )
                if svar.upper()[0] == "J":
                    True
                else:
                    break
        else:
            print("Du må bestille minst 1 billett")
    else:
        True
wb.save("LedigDato.xlsx")


# Lager billetten
ny_billett = Billett(forestilling, dato, honnor, voksen, student, barn)
ny_billett.eksporter_bestilling(epost)

wb = openpyxl.load_workbook(epost + ".xlsx")
sheet = wb.active


ny_billett.vis_kvittering()
